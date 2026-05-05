import json
from pathlib import Path
from xml.etree import ElementTree as ET
from zipfile import ZipFile

import openpyxl
import pandas as pd
import pytest

from api.app.report_models import (
    ChartSpec,
    KpiCard,
    NarrativeBlock,
    ReportMeta,
    ReportSpec,
    TemplateDecision,
)
from api.app.services.analytics import compute_analysis
from api.app.services.charts import generate_chart_bundle
from api.app.services.excel import render_report


CHART_TITLES = {
    "histogram": "测量值分布图",
    "control_chart_imr": "I-MR 控制图",
    "trend_line": "测量值趋势图",
    "spec_comparison": "测量值与规格对比图",
}


def _build_report_spec(template_id: str, chart_ids: list[str]) -> ReportSpec:
    return ReportSpec(
        report_meta=ReportMeta(
            title="轴承检测 SPC 报告",
            report_id="RPT-002",
            generated_at="2026-04-21T10:00:00Z",
            batch_id="B-02",
            product_name="6205 轴承",
        ),
        template_decision=TemplateDecision(
            template_id=template_id,
            reason="本批次适合使用总览模板展示",
        ),
        dataset_summary={"sample_count": 6, "overall_pass_rate": 0.8333},
        kpi_cards=[
            KpiCard(label="均值", value="10.022"),
            KpiCard(label="Cpk", value="0.91"),
        ],
        detail_rows=[
            {"sample_id": "A-001", "measurement_value": 10.01, "status": "PASS"},
            {"sample_id": "A-002", "measurement_value": 10.06, "status": "FAIL"},
        ],
        chart_specs=[
            ChartSpec(chart_id=chart_id, chart_type=chart_id, title=CHART_TITLES[chart_id])
            for chart_id in chart_ids
        ],
        anomalies=[{"severity": "high", "summary": "1 point exceeds spec"}],
        ai_narrative=NarrativeBlock(
            executive_summary="过程中心呈上升漂移趋势。",
            quality_risk="过程能力偏弱，建议尽快复核。",
            recommended_actions=["检查设备漂移", "复核最近一次调机变更"],
        ),
    )


def _build_chart_paths(tmp_path: Path, chart_count: int) -> dict[str, str]:
    normalized = pd.DataFrame(
        {
            "measurement_value": [10.01, 10.02, 10.03, 10.00, 10.06, 10.01],
            "usl": [10.05] * 6,
            "lsl": [9.95] * 6,
            "sequence_index": [1, 2, 3, 4, 5, 6],
        }
    )
    analysis = compute_analysis(normalized)
    chart_paths = generate_chart_bundle(normalized, analysis, tmp_path)
    selected_chart_ids = list(chart_paths.keys())[:chart_count]
    return {chart_id: chart_paths[chart_id] for chart_id in selected_chart_ids}


def _write_template_manifest(
    tmp_path: Path,
    include_chart_layout: bool,
    include_formal_layout: bool = False,
) -> Path:
    templates_root = tmp_path / "templates"
    template_dir = templates_root / "test_template"
    template_dir.mkdir(parents=True)

    slots: dict[str, object] = {
        "title_cell": "A1",
        "product_cell": "A3",
        "batch_cell": "A4",
        "reason_cell": "A6",
        "kpi_start_cell": "A8",
        "narrative_summary_cell": "A14",
        "narrative_risk_cell": "A15",
        "actions_start_cell": "A17",
        "detail_start_row": 2,
    }
    if include_formal_layout:
        slots.update(
            {
                "product_cell": "B3",
                "batch_cell": "B4",
                "reason_cell": "B5",
                "narrative_summary_cell": "A15",
                "narrative_risk_cell": "A21",
                "actions_start_cell": "A26",
            }
        )

    manifest: dict[str, object] = {
        "template_id": "test_template",
        "workbook_name": "test_template.xlsx",
        "summary_sheet": "Summary",
        "detail_sheet": "Details",
        "slots": slots,
    }
    if include_chart_layout:
        manifest["chart_layout"] = {
            "sheet": "Summary",
            "area_title_cell": "H1",
            "area_title_text": "SPC 图表总览",
            "layouts": {
                "1": [
                    {"title_cell": "H2", "image_cell": "H3", "width": 640, "height": 360}
                ],
                "2": [
                    {"title_cell": "H2", "image_cell": "H3", "width": 640, "height": 170},
                    {"title_cell": "H21", "image_cell": "H22", "width": 640, "height": 170},
                ],
                "3": [
                    {"title_cell": "H2", "image_cell": "H3", "width": 300, "height": 170},
                    {"title_cell": "N2", "image_cell": "N3", "width": 300, "height": 170},
                    {"title_cell": "K21", "image_cell": "K22", "width": 420, "height": 170},
                ],
                "4": [
                    {"title_cell": "H2", "image_cell": "H3", "width": 300, "height": 170},
                    {"title_cell": "N2", "image_cell": "N3", "width": 300, "height": 170},
                    {"title_cell": "H21", "image_cell": "H22", "width": 300, "height": 170},
                    {"title_cell": "N21", "image_cell": "N22", "width": 300, "height": 170},
                ],
            },
        }
    if include_formal_layout:
        manifest["formal_layout"] = {
            "column_widths": {
                "A": 16,
                "B": 18,
                "C": 14,
                "D": 2.5,
                "H": 18,
                "I": 18,
                "J": 18,
                "K": 18,
                "L": 18,
                "M": 18,
                "N": 18,
                "O": 18,
                "P": 18,
                "Q": 18,
            },
            "title_band": {"range": "A1:D1"},
            "metadata_rows": [
                {"label_cell": "A3", "value_cell": "B3", "label": "产品"},
                {"label_cell": "A4", "value_cell": "B4", "label": "批次"},
                {"label_cell": "A5", "value_cell": "B5", "label": "选型依据"},
            ],
            "sections": {
                "kpi": {
                    "header_cell": "A7",
                    "header_text": "关键指标",
                    "body_range": "A8:B12",
                    "kind": "kpi",
                },
                "summary": {
                    "header_cell": "A14",
                    "header_text": "分析结论",
                    "body_range": "A15:D18",
                    "kind": "text",
                },
                "risk": {
                    "header_cell": "A20",
                    "header_text": "质量风险",
                    "body_range": "A21:D23",
                    "kind": "text",
                },
                "actions": {
                    "header_cell": "A25",
                    "header_text": "建议措施",
                    "body_range": "A26:D29",
                    "kind": "actions",
                },
            },
            "chart_board": {
                "title_cell": "H1",
                "title_text": "SPC 图表总览",
                "frames": {
                    "1": ["H2:Q20"],
                    "2": ["H2:Q19", "H21:Q38"],
                    "3": ["H2:M19", "N2:Q19", "K21:P38"],
                    "4": ["H2:M19", "N2:Q19", "H21:M38", "N21:Q38"],
                },
            },
        }

    (template_dir / "template_manifest.json").write_text(
        json.dumps(manifest, indent=2),
        encoding="utf-8",
    )
    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    workbook.create_sheet("Details")
    workbook.save(template_dir / "test_template.xlsx")
    return templates_root


def _read_drawing_anchors(report_file: Path) -> list[tuple[int, int, int, int]]:
    with ZipFile(report_file) as archive:
        drawing_xml = archive.read("xl/drawings/drawing1.xml")

    root = ET.fromstring(drawing_xml)
    namespace = {"x": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"}
    anchors: list[tuple[int, int, int, int]] = []
    for one_cell_anchor in root.findall("x:oneCellAnchor", namespace):
        from_node = one_cell_anchor.find("x:from", namespace)
        ext_node = one_cell_anchor.find("x:ext", namespace)
        if from_node is None or ext_node is None:
            continue
        col_node = from_node.find("x:col", namespace)
        row_node = from_node.find("x:row", namespace)
        if col_node is None or row_node is None:
            continue
        anchors.append(
            (
                int(col_node.text),
                int(row_node.text),
                int(ext_node.attrib["cx"]),
                int(ext_node.attrib["cy"]),
            )
        )
    return anchors


def test_render_report_populates_template(tmp_path: Path) -> None:
    output_path = tmp_path / "demo_report.xlsx"
    report_spec = ReportSpec(
        report_meta=ReportMeta(
            title="轴承检测 SPC 报告",
            report_id="RPT-001",
            generated_at="2026-04-21T10:00:00Z",
            batch_id="B-01",
            product_name="6205 轴承",
        ),
        template_decision=TemplateDecision(
            template_id="template_a_overview",
            reason="本批次适合使用总览模板展示",
        ),
        dataset_summary={"sample_count": 6, "overall_pass_rate": 0.8333},
        kpi_cards=[
            KpiCard(label="均值", value="10.022"),
            KpiCard(label="Cpk", value="0.91"),
        ],
        detail_rows=[
            {"sample_id": "A-001", "measurement_value": 10.01, "status": "PASS"},
            {"sample_id": "A-002", "measurement_value": 10.06, "status": "FAIL"},
        ],
        chart_specs=[ChartSpec(chart_id="histogram", chart_type="histogram", title="测量值分布图")],
        anomalies=[{"severity": "high", "summary": "1 point exceeds spec"}],
        ai_narrative=NarrativeBlock(
            executive_summary="过程中心呈上升漂移趋势。",
            quality_risk="过程能力偏弱，建议尽快复核。",
            recommended_actions=["检查设备漂移", "复核最近一次调机变更"],
        ),
    )

    report_file = render_report(
        report_spec=report_spec,
        chart_paths={},
        templates_root=Path("templates"),
        output_path=output_path,
    )

    workbook = openpyxl.load_workbook(report_file)
    summary_sheet = workbook["Summary"]

    assert report_file.exists()
    assert summary_sheet["A1"].value == "轴承检测 SPC 报告"
    assert summary_sheet["B4"].value == "B-01"
    assert summary_sheet["A7"].value == "关键指标"
    assert summary_sheet["A8"].value == "均值"
    assert summary_sheet["A15"].value == "过程中心呈上升漂移趋势。"
    assert summary_sheet["H1"].value == "SPC 图表总览"


@pytest.mark.parametrize(
    ("chart_count", "expected_title_cells"),
    [
        (1, ["H2"]),
        (2, ["H2", "H21"]),
        (3, ["H2", "N2", "K21"]),
        (4, ["H2", "N2", "H21", "N21"]),
    ],
)
def test_render_report_places_chart_images_on_summary_sheet(
    tmp_path: Path,
    chart_count: int,
    expected_title_cells: list[str],
) -> None:
    templates_root = _write_template_manifest(tmp_path, include_chart_layout=True)
    chart_paths = _build_chart_paths(tmp_path / f"charts-{chart_count}", chart_count)
    report_file = render_report(
        report_spec=_build_report_spec("test_template", list(chart_paths.keys())),
        chart_paths=chart_paths,
        templates_root=templates_root,
        output_path=tmp_path / f"summary-layout-{chart_count}.xlsx",
    )

    workbook = openpyxl.load_workbook(report_file)
    summary_sheet = workbook["Summary"]

    assert "Charts" not in workbook.sheetnames
    assert len(getattr(summary_sheet, "_images", [])) == chart_count
    for title_cell, chart_id in zip(expected_title_cells, chart_paths):
        assert summary_sheet[title_cell].value == CHART_TITLES[chart_id]


def test_render_report_requires_chart_layout_for_chart_embedding(tmp_path: Path) -> None:
    templates_root = _write_template_manifest(tmp_path, include_chart_layout=False)
    chart_paths = _build_chart_paths(tmp_path / "charts-fallback", 4)
    with pytest.raises(ValueError, match="chart_layout"):
        render_report(
            report_spec=_build_report_spec("test_template", list(chart_paths.keys())),
            chart_paths=chart_paths,
            templates_root=templates_root,
            output_path=tmp_path / "fallback-layout.xlsx",
        )


def test_render_report_requires_existing_template_workbook(tmp_path: Path) -> None:
    templates_root = _write_template_manifest(tmp_path, include_chart_layout=True)
    workbook_path = templates_root / "test_template" / "test_template.xlsx"
    workbook_path.unlink()

    with pytest.raises(FileNotFoundError, match="test_template.xlsx"):
        render_report(
            report_spec=_build_report_spec("test_template", []),
            chart_paths={},
            templates_root=templates_root,
            output_path=tmp_path / "missing-template.xlsx",
        )



def test_render_report_applies_formal_summary_style_on_summary_sheet(tmp_path: Path) -> None:
    templates_root = _write_template_manifest(
        tmp_path,
        include_chart_layout=True,
        include_formal_layout=True,
    )
    chart_paths = _build_chart_paths(tmp_path / "charts-formal", 4)
    report_file = render_report(
        report_spec=_build_report_spec("test_template", list(chart_paths.keys())),
        chart_paths=chart_paths,
        templates_root=templates_root,
        output_path=tmp_path / "formal-summary.xlsx",
    )

    workbook = openpyxl.load_workbook(report_file)
    summary_sheet = workbook["Summary"]
    merged_ranges = {str(cell_range) for cell_range in summary_sheet.merged_cells.ranges}

    assert "A1:D1" in merged_ranges
    assert summary_sheet["A1"].fill.fgColor.rgb == "FF1F2937"
    assert summary_sheet["A1"].font.bold is True
    assert summary_sheet["A3"].value == "产品"
    assert summary_sheet["B3"].value == "6205 轴承"
    assert summary_sheet["A7"].value == "关键指标"
    assert summary_sheet["A8"].border.left.style == "thin"
    assert summary_sheet["B8"].font.bold is True
    assert summary_sheet["A14"].value == "分析结论"
    assert summary_sheet["A15"].value == "过程中心呈上升漂移趋势。"
    assert summary_sheet["A20"].value == "质量风险"
    assert summary_sheet["A21"].value == "过程能力偏弱，建议尽快复核。"
    assert summary_sheet["A25"].value == "建议措施"
    assert summary_sheet["A26"].value == "- 检查设备漂移"
    assert summary_sheet["H1"].value == "SPC 图表总览"
    assert summary_sheet["H2"].font.bold is True
    assert summary_sheet["H2"].border.top.style == "thin"


def test_render_report_keeps_plain_summary_when_formal_layout_is_missing(tmp_path: Path) -> None:
    templates_root = _write_template_manifest(
        tmp_path,
        include_chart_layout=True,
        include_formal_layout=False,
    )
    chart_paths = _build_chart_paths(tmp_path / "charts-no-formal", 2)
    report_file = render_report(
        report_spec=_build_report_spec("test_template", list(chart_paths.keys())),
        chart_paths=chart_paths,
        templates_root=templates_root,
        output_path=tmp_path / "plain-summary.xlsx",
    )

    workbook = openpyxl.load_workbook(report_file)
    summary_sheet = workbook["Summary"]

    assert "A1:D1" not in {str(cell_range) for cell_range in summary_sheet.merged_cells.ranges}
    assert summary_sheet["A1"].fill.patternType is None
    assert summary_sheet["A8"].value == "均值"
    assert summary_sheet["H1"].value == "SPC 图表总览"
    assert summary_sheet["H1"].fill.patternType is None

from api.app.services.templates import load_template_manifest


def test_demo_template_manifests_define_summary_chart_layout() -> None:
    expected_title_cells = {
        "template_a_overview": "H1",
        "template_b_detailed": "J1",
        "template_c_showcase": "J2",
    }
    for template_id, title_cell in expected_title_cells.items():
        manifest = load_template_manifest(Path("templates"), template_id)
        chart_layout = manifest["chart_layout"]
        assert chart_layout["sheet"] == "Summary"
        assert chart_layout["area_title_cell"] == title_cell
        assert set(chart_layout["layouts"].keys()) == {"1", "2", "3", "4"}


def test_demo_template_manifests_define_denser_three_and_four_chart_layouts() -> None:
    expected_layouts = {
        "template_a_overview": {
            "3": [
                {"title_cell": "H2", "image_cell": "H3", "width": 440, "height": 220},
                {"title_cell": "N2", "image_cell": "N3", "width": 440, "height": 220},
                {"title_cell": "K14", "image_cell": "K15", "width": 620, "height": 220},
            ],
            "4": [
                {"title_cell": "H2", "image_cell": "H3", "width": 560, "height": 220},
                {"title_cell": "M2", "image_cell": "M3", "width": 560, "height": 220},
                {"title_cell": "H14", "image_cell": "H15", "width": 560, "height": 220},
                {"title_cell": "M14", "image_cell": "M15", "width": 560, "height": 220},
            ],
            "frames": {
                "3": ["H2:M13", "N2:Q13", "K14:P25"],
                "4": ["H2:L25", "M2:Q25"],
            },
        },
        "template_b_detailed": {
            "3": [
                {"title_cell": "J2", "image_cell": "J3", "width": 440, "height": 220},
                {"title_cell": "P2", "image_cell": "P3", "width": 440, "height": 220},
                {"title_cell": "M14", "image_cell": "M15", "width": 620, "height": 220},
            ],
            "4": [
                {"title_cell": "J2", "image_cell": "J3", "width": 440, "height": 220},
                {"title_cell": "N2", "image_cell": "N3", "width": 440, "height": 220},
                {"title_cell": "J14", "image_cell": "J15", "width": 440, "height": 220},
                {"title_cell": "N14", "image_cell": "N15", "width": 440, "height": 220},
            ],
            "frames": {
                "3": ["J2:M13", "P2:Q13", "M14:Q25"],
                "4": ["J2:M25", "N2:Q25"],
            },
        },
        "template_c_showcase": {
            "3": [
                {"title_cell": "J3", "image_cell": "J4", "width": 440, "height": 220},
                {"title_cell": "P3", "image_cell": "P4", "width": 440, "height": 220},
                {"title_cell": "M15", "image_cell": "M16", "width": 620, "height": 220},
            ],
            "4": [
                {"title_cell": "J3", "image_cell": "J4", "width": 440, "height": 220},
                {"title_cell": "N3", "image_cell": "N4", "width": 440, "height": 220},
                {"title_cell": "J15", "image_cell": "J16", "width": 440, "height": 220},
                {"title_cell": "N15", "image_cell": "N16", "width": 440, "height": 220},
            ],
            "frames": {
                "3": ["J3:M14", "P3:Q14", "M15:Q26"],
                "4": ["J3:M26", "N3:Q26"],
            },
        },
    }

    for template_id, expected in expected_layouts.items():
        manifest = load_template_manifest(Path("templates"), template_id)
        chart_layout = manifest["chart_layout"]
        formal_layout = manifest["formal_layout"]
        assert chart_layout["layouts"]["3"] == expected["3"]
        assert chart_layout["layouts"]["4"] == expected["4"]
        assert formal_layout["chart_board"]["frames"]["3"] == expected["frames"]["3"]
        assert formal_layout["chart_board"]["frames"]["4"] == expected["frames"]["4"]



def test_demo_template_manifests_define_formal_summary_layout() -> None:
    expected_layout = {
        "template_a_overview": {
            "title_cell": "A1",
            "kpi_start_cell": "A8",
            "chart_title_cell": "H1",
        },
        "template_b_detailed": {
            "title_cell": "A1",
            "kpi_start_cell": "A8",
            "chart_title_cell": "J1",
        },
        "template_c_showcase": {
            "title_cell": "B2",
            "kpi_start_cell": "B9",
            "chart_title_cell": "J2",
        },
    }
    for template_id, expected in expected_layout.items():
        manifest = load_template_manifest(Path("templates"), template_id)
        formal_layout = manifest["formal_layout"]
        assert manifest["slots"]["title_cell"] == expected["title_cell"]
        assert manifest["slots"]["kpi_start_cell"] == expected["kpi_start_cell"]
        assert formal_layout["sections"]["kpi"]["header_text"] == "关键指标"
        assert formal_layout["chart_board"]["title_cell"] == expected["chart_title_cell"]
        assert set(formal_layout["chart_board"]["frames"].keys()) == {"1", "2", "3", "4"}


@pytest.mark.parametrize(
    ("chart_count", "expected_title_cells"),
    [
        (3, ["H2", "N2", "K14"]),
        (4, ["H2", "M2", "H14", "M14"]),
    ],
)
def test_render_report_places_generated_charts_on_demo_summary_sheet(
    tmp_path: Path,
    chart_count: int,
    expected_title_cells: list[str],
) -> None:
    chart_paths = _build_chart_paths(tmp_path / f"charts-demo-{chart_count}", chart_count)
    report_file = render_report(
        report_spec=_build_report_spec("template_a_overview", list(chart_paths.keys())),
        chart_paths=chart_paths,
        templates_root=Path("templates"),
        output_path=tmp_path / f"demo-summary-layout-{chart_count}.xlsx",
    )

    workbook = openpyxl.load_workbook(report_file)
    summary_sheet = workbook["Summary"]

    assert "Charts" not in workbook.sheetnames
    assert len(getattr(summary_sheet, "_images", [])) == chart_count
    assert "A1:D1" in {str(cell_range) for cell_range in summary_sheet.merged_cells.ranges}
    assert summary_sheet["A7"].value == "关键指标"
    assert summary_sheet["H1"].value == "SPC 图表总览"
    for title_cell, chart_id in zip(expected_title_cells, chart_paths):
        assert summary_sheet[title_cell].value == CHART_TITLES[chart_id]


def test_render_report_uses_tighter_horizontal_spacing_for_four_demo_charts(tmp_path: Path) -> None:
    chart_paths = _build_chart_paths(tmp_path / "charts-demo-horizontal-4", 4)
    report_file = render_report(
        report_spec=_build_report_spec("template_a_overview", list(chart_paths.keys())),
        chart_paths=chart_paths,
        templates_root=Path("templates"),
        output_path=tmp_path / "demo-summary-horizontal-4.xlsx",
    )

    assert _read_drawing_anchors(report_file) == [
        (7, 2, 5334000, 2095500),
        (12, 2, 5334000, 2095500),
        (7, 14, 5334000, 2095500),
        (12, 14, 5334000, 2095500),
    ]
