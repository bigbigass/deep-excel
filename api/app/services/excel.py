"""Excel 报表渲染服务。

这个模块不负责“决定报表写什么”，只负责“把已经决定好的内容写进模板”。
换句话说：
- `ReportSpec` 决定内容；
- `template_manifest.json` 决定布局；
- 本模块负责把两者拼在一起并输出 Excel。
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet

from api.app.report_models import ReportSpec
from api.app.services.excel_styles import apply_formal_summary_style
from api.app.services.report_localization import chart_title
from api.app.services.templates import ensure_demo_template_workbook, load_template_manifest


def _iter_chart_images(report_spec: ReportSpec, chart_paths: dict[str, str]) -> list[tuple[str, str]]:
    """按照报表定义的顺序挑选需要嵌入的图表，最多保留四张。

    先按 `report_spec.chart_specs` 的顺序取图，保证 AI 规划的展示顺序优先；
    如果 chart_paths 里还有未被声明的图，再按兜底顺序追加。
    """
    ordered_charts: list[tuple[str, str]] = []
    seen_chart_ids: set[str] = set()

    for chart_spec in report_spec.chart_specs:
        chart_path = chart_paths.get(chart_spec.chart_id)
        if chart_path is None:
            continue
        ordered_charts.append((chart_spec.title, chart_path))
        seen_chart_ids.add(chart_spec.chart_id)

    for chart_id, chart_path in chart_paths.items():
        if chart_id in seen_chart_ids:
            continue
        ordered_charts.append((chart_title(chart_id), chart_path))

    return ordered_charts[:4]


def _embed_chart_images_on_summary_sheet(
    workbook: openpyxl.Workbook,
    summary_sheet: Worksheet,
    chart_images: list[tuple[str, str]],
    chart_layout: dict[str, object],
) -> None:
    """根据模板中的布局定义，把图表图片放到摘要页。

    这里完全依赖 manifest 声明的位置和尺寸，不在代码里写死单元格，
    这样换模板时只需要改模板配置，不用改 Python 逻辑。
    """
    layouts = chart_layout.get("layouts")
    if not isinstance(layouts, dict):
        raise ValueError("chart_layout.layouts must be a mapping")

    placements = layouts.get(str(len(chart_images)))
    if not isinstance(placements, list) or len(placements) < len(chart_images):
        raise ValueError("chart_layout must define placements for every chart image")

    layout_sheet_name = str(chart_layout.get("sheet", summary_sheet.title))
    if layout_sheet_name not in workbook.sheetnames:
        raise ValueError(f"chart_layout references missing sheet: {layout_sheet_name}")

    target_sheet = workbook[layout_sheet_name]
    target_sheet[str(chart_layout.get("area_title_cell", "H1"))] = str(
        chart_layout.get("area_title_text", "SPC 图表总览")
    )

    for (title, chart_path), placement in zip(chart_images, placements):
        if not isinstance(placement, dict):
            raise ValueError("chart_layout placements must be objects")
        image = Image(chart_path)
        image.width = int(placement["width"])
        image.height = int(placement["height"])
        target_sheet[str(placement["title_cell"])] = title
        target_sheet.add_image(image, str(placement["image_cell"]))


def _embed_chart_images(
    workbook: openpyxl.Workbook,
    summary_sheet: Worksheet,
    report_spec: ReportSpec,
    chart_paths: dict[str, str],
    manifest: dict[str, object],
) -> None:
    """在模板声明支持图表布局时执行图表嵌入。

    图表是可选能力，所以这里先判断有没有图片、模板有没有布局配置，
    再决定是否执行嵌入。
    """
    chart_images = _iter_chart_images(report_spec, chart_paths)
    if not chart_images:
        return

    chart_layout = manifest.get("chart_layout")
    if not isinstance(chart_layout, dict):
        raise ValueError("template manifest must define chart_layout for chart embedding")

    _embed_chart_images_on_summary_sheet(
        workbook=workbook,
        summary_sheet=summary_sheet,
        chart_images=chart_images,
        chart_layout=chart_layout,
    )


def render_report(
    report_spec: ReportSpec,
    chart_paths: dict[str, str],
    templates_root: Path,
    output_path: Path,
) -> Path:
    """把 AI 规划结果和图表资源写入 Excel 模板并导出成最终报表。

    整体分三步：
    1. 打开模板和 manifest，定位摘要页/明细页/插槽位置。
    2. 把 KPI、叙述、明细和图表写入指定位置。
    3. 套用样式并保存到 outputs。
    """
    template_id = report_spec.template_decision.template_id
    workbook_path = ensure_demo_template_workbook(templates_root, template_id)
    manifest = load_template_manifest(templates_root, template_id)
    workbook = openpyxl.load_workbook(workbook_path)
    summary_sheet = workbook[str(manifest["summary_sheet"])]
    detail_sheet = workbook[str(manifest["detail_sheet"])]
    slots = manifest["slots"]

    summary_sheet[str(slots["title_cell"])] = report_spec.report_meta.title
    summary_sheet[str(slots["product_cell"])] = report_spec.report_meta.product_name
    summary_sheet[str(slots["batch_cell"])] = report_spec.report_meta.batch_id
    summary_sheet[str(slots["reason_cell"])] = report_spec.template_decision.reason

    # summary 页的文本区域都由 manifest 提供坐标，代码只负责按槽位填值。
    kpi_cell = summary_sheet[str(slots["kpi_start_cell"])]
    start_row = kpi_cell.row
    start_col = kpi_cell.column
    # KPI 从模板声明的起始单元格向下逐行填充，保持模板可配置。
    for offset, card in enumerate(report_spec.kpi_cards):
        summary_sheet.cell(row=start_row + offset, column=start_col, value=card.label)
        summary_sheet.cell(row=start_row + offset, column=start_col + 1, value=card.value)

    summary_sheet[str(slots["narrative_summary_cell"])] = report_spec.ai_narrative.executive_summary
    summary_sheet[str(slots["narrative_risk_cell"])] = report_spec.ai_narrative.quality_risk
    action_row = summary_sheet[str(slots["actions_start_cell"])].row
    action_col = summary_sheet[str(slots["actions_start_cell"])].column
    # 动作建议按行展开，模板只需要给出起始位置。
    for offset, action in enumerate(report_spec.ai_narrative.recommended_actions):
        summary_sheet.cell(row=action_row + offset, column=action_col, value=f"- {action}")

    detail_start_row = int(slots["detail_start_row"])
    # 当前明细页只写最关键的三列，保持 demo 模板简单直观。
    for row_offset, detail_row in enumerate(report_spec.detail_rows):
        detail_sheet.cell(row=detail_start_row + row_offset, column=1, value=detail_row.get("sample_id"))
        detail_sheet.cell(row=detail_start_row + row_offset, column=2, value=detail_row.get("measurement_value"))
        detail_sheet.cell(row=detail_start_row + row_offset, column=3, value=detail_row.get("status"))

    _embed_chart_images(
        workbook=workbook,
        summary_sheet=summary_sheet,
        report_spec=report_spec,
        chart_paths=chart_paths,
        manifest=manifest,
    )
    apply_formal_summary_style(summary_sheet=summary_sheet, manifest=manifest)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
