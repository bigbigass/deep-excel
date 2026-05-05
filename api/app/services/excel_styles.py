"""Excel 摘要页样式工具。"""

from __future__ import annotations

from collections.abc import Mapping

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

THIN_SIDE = Side(style="thin", color="FFCBD5E1")
BLOCK_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
TITLE_FILL = PatternFill(fill_type="solid", fgColor="FF1F2937")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="FF334155")
LABEL_FILL = PatternFill(fill_type="solid", fgColor="FFF8FAFC")
CHART_TITLE_FILL = PatternFill(fill_type="solid", fgColor="FFE2E8F0")


def apply_formal_summary_style(summary_sheet: Worksheet, manifest: dict[str, object]) -> None:
    """根据模板 manifest 为正式摘要页套用统一视觉样式。"""
    formal_layout = manifest.get("formal_layout")
    if not isinstance(formal_layout, Mapping):
        return

    _apply_column_widths(summary_sheet, formal_layout.get("column_widths", {}))
    _style_title_band(summary_sheet, formal_layout.get("title_band", {}))
    _style_metadata_rows(summary_sheet, formal_layout.get("metadata_rows", []))

    sections = formal_layout.get("sections", {})
    if isinstance(sections, Mapping):
        for section in sections.values():
            _style_section(summary_sheet, section)

    _style_chart_board(summary_sheet, manifest, formal_layout.get("chart_board", {}))


def _apply_column_widths(summary_sheet: Worksheet, column_widths: object) -> None:
    """批量设置列宽。"""
    if not isinstance(column_widths, Mapping):
        return

    for column_letter, width in column_widths.items():
        summary_sheet.column_dimensions[str(column_letter)].width = float(width)


def _style_title_band(summary_sheet: Worksheet, title_band: object) -> None:
    """设置报表标题横幅。"""
    if not isinstance(title_band, Mapping) or "range" not in title_band:
        return

    cell_range = str(title_band["range"])
    summary_sheet.merge_cells(cell_range)
    title_cell = summary_sheet[cell_range.split(":", maxsplit=1)[0]]
    title_cell.fill = TITLE_FILL
    title_cell.font = Font(size=16, bold=True, color="FFFFFFFF")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.border = BLOCK_BORDER


def _style_metadata_rows(summary_sheet: Worksheet, metadata_rows: object) -> None:
    """渲染标题下方的元数据标签和值。"""
    if not isinstance(metadata_rows, list):
        return

    for metadata_row in metadata_rows:
        if not isinstance(metadata_row, Mapping):
            continue
        label_cell = summary_sheet[str(metadata_row["label_cell"])]
        value_cell = summary_sheet[str(metadata_row["value_cell"])]
        label_cell.value = str(metadata_row["label"])
        label_cell.font = Font(bold=True, color="FF0F172A")
        label_cell.fill = LABEL_FILL
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        label_cell.border = BLOCK_BORDER
        value_cell.alignment = Alignment(horizontal="left", vertical="center")
        value_cell.border = BLOCK_BORDER


def _style_section(summary_sheet: Worksheet, section: object) -> None:
    """统一处理 KPI、文本区块等摘要页分区样式。"""
    if not isinstance(section, Mapping):
        return

    header_cell = summary_sheet[str(section["header_cell"])]
    header_cell.value = str(section["header_text"])
    header_cell.fill = SECTION_FILL
    header_cell.font = Font(bold=True, color="FFFFFFFF")
    header_cell.alignment = Alignment(horizontal="left", vertical="center")
    header_cell.border = BLOCK_BORDER

    body_range = str(section["body_range"])
    min_col, min_row, max_col, max_row = range_boundaries(body_range)
    for row_index in range(min_row, max_row + 1):
        for column_index in range(min_col, max_col + 1):
            summary_sheet.cell(row=row_index, column=column_index).border = BLOCK_BORDER

    kind = str(section.get("kind", "text"))
    if kind == "kpi":
        for row_index in range(min_row, max_row + 1):
            label_cell = summary_sheet.cell(row=row_index, column=min_col)
            value_cell = summary_sheet.cell(row=row_index, column=max_col)
            label_cell.fill = LABEL_FILL
            label_cell.alignment = Alignment(horizontal="left", vertical="center")
            value_cell.font = Font(bold=True, color="FF0F172A")
            value_cell.alignment = Alignment(horizontal="right", vertical="center")
        return

    for row_index in range(min_row, max_row + 1):
        for column_index in range(min_col, max_col + 1):
            summary_sheet.cell(row=row_index, column=column_index).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )


def _style_chart_board(summary_sheet: Worksheet, manifest: dict[str, object], chart_board: object) -> None:
    """为图表总览区补标题和边框，和图片布局保持一致。"""
    if not isinstance(chart_board, Mapping):
        return

    title_cell = summary_sheet[str(chart_board.get("title_cell", "H1"))]
    title_cell.value = str(chart_board.get("title_text", title_cell.value or "SPC 图表总览"))
    title_cell.fill = SECTION_FILL
    title_cell.font = Font(size=12, bold=True, color="FFFFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = BLOCK_BORDER

    chart_layout = manifest.get("chart_layout")
    placements: list[object] = []
    if isinstance(chart_layout, Mapping):
        layouts = chart_layout.get("layouts")
        chart_count = len(getattr(summary_sheet, "_images", []))
        if isinstance(layouts, Mapping):
            selected = layouts.get(str(chart_count), [])
            if isinstance(selected, list):
                placements = selected

        for placement in placements:
            if not isinstance(placement, Mapping):
                continue
            chart_title = summary_sheet[str(placement["title_cell"])]
            chart_title.fill = CHART_TITLE_FILL
            chart_title.font = Font(size=10, bold=True, color="FF0F172A")
            chart_title.alignment = Alignment(horizontal="left", vertical="center")
            chart_title.border = BLOCK_BORDER

    frames = chart_board.get("frames")
    chart_count = len(getattr(summary_sheet, "_images", []))
    if isinstance(frames, Mapping):
        selected_frames = frames.get(str(chart_count), [])
        if isinstance(selected_frames, list):
            for frame_range in selected_frames:
                if not isinstance(frame_range, str):
                    continue
                _apply_border_to_range(summary_sheet, frame_range)


def _apply_border_to_range(summary_sheet: Worksheet, cell_range: str) -> None:
    """给连续区域补边框，方便在模板中形成图表卡片。"""
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row_index in range(min_row, max_row + 1):
        for column_index in range(min_col, max_col + 1):
            summary_sheet.cell(row=row_index, column=column_index).border = BLOCK_BORDER
